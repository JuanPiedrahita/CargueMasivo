import openpyxl
from optparse import OptionParser
from bdConnection import Connection

def main():
    parser = OptionParser()

    parser.add_option("-N", "--db_name", dest="db_name", help="database name", default="financiera")
    parser.add_option("-U", "--db_user",dest="db_user",help="database user", default="postgres")
    parser.add_option("-P", "--db_password", dest="db_password", help="database password", default="postgres")
    parser.add_option("-H", "--host_serverBD", dest="host_serverBD", help="server host", default="localhost")
    parser.add_option("-K", "--port_serverBD", dest="port_serverBD", help="server port", default="5432")
    parser.add_option("-p", "--path_csv", dest="path_csv", help="path of file for uploading", default="csv/concepto.csv")
    parser.add_option("-d", "--debug", dest="debug", help="Mostrar mensajes de debug utilize 10", default=10)

    (options, args) = parser.parse_args()

    if not options.db_name:
        parser.error('Parametro db_name no especificado')
    if not options.db_user:
        parser.error('Parametro db_user no especificado') 
    if not options.db_password:
        parser.error('Parametro db_password no especificado')
    if not options.host_serverBD:
        parser.error('Parametro host_serverBD no especificado')
    if not options.port_serverBD:
        parser.error('Parametro host_serverBD no especificado')
    if not options.path_csv:
        parser.error('Parametro path_csv no especificado')

    connect = Connection(options)
    postgres_connect = connect.get_connection()
    cursor = connect.get_cursor()




if __name__ == '__main__':
	main()

#doc = openpyxl.load_workbook('archivosExcel/cuentas.xlsx')

#hojaData = doc.get_sheet_by_name('HojaCuentas')
#hojaErrores = doc.create_sheet('HojaErrores')
#hojaCompletos = doc.create_sheet('HojaCompletos')



#f = 1
#for fila in hojaData.rows:
#	c = 1
#	for columna in fila:
#		hojaErrores.cell(row=f, column=c).value = columna.value
#		c=c+1
#		print columna.value
#	f=f+1
#	print ""


#hojaErrores.append(["hola","hola"])

#doc.save("archivosExcel/resultadoCargue.xlsx")

