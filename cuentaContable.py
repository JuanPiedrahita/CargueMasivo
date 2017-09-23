import openpyxl
from cuenta import Cuenta

class CuentaContable():
	def __init__(self, cursor, fileAddress):
		self.cursor = cursor
		self.fileAddress = fileAddress

	def cargarCuentas(self, ruta):
		doc = openpyxl.load_workbook(self.fileAddress)
		hojaData = doc.get_sheet_by_name('HojaCuentas')
		hojaErrores = doc.create_sheet('HojaErrores')
		hojaCompletos = doc.create_sheet('HojaCompletos')
		hojaCompletos.append(['codigo', 'nombre', 'padre', 'naturaleza', 'nivel', 'idCuenta', 'idPadre', 'idRelacion'])
		hojaErrores.append(['codigo' , 'errores', 'nombre'])

		rowCounter = 1
		for fila in hojaData.rows:
			code = self.cleanIdCuenta(hojaData.cell(row=rowCounter,column=1).value)
			nombreCuenta = hojaData.cell(row=rowCounter,column=2).value
			rowCounter += 1
			print "Evaluando cuenta "+code	
			cuenta = Cuenta(code, nombreCuenta)
			cuenta.padre = cuenta.definePadre(cuenta.codigo)
			cuenta.idCuenta = self.getCuentaById(cuenta.codigo)
			cuenta.idPadre = self.getCuentaById(cuenta.padre)
			cuenta.naturaleza = cuenta.defineNaturaleza(cuenta.codigo[:1])
			cuenta.nivel = cuenta.defineNivel(cuenta.codigo)
			error = []
			if cuenta.idCuenta:
				error.append("La cuenta ya existe.")
			if len(cuenta.nombre.replace("'",""))<3:
				error.append("El nombre de la cuenta no es valido.")
			if cuenta.naturaleza is None:
				error.append("La naturaleza de la cuenta no esta definida.")
			if not cuenta.idPadre:
				error.append("El padre no se encuentra registrado.")

			if not error:
				self.insertCuenta(cuenta)
				hojaCompletos.append([cuenta.codigo, cuenta.nombre, cuenta.padre, cuenta.naturaleza, cuenta.nivel, cuenta.idCuenta, cuenta.idPadre, cuenta.idRelacion])
			else:	
				print "Error: ", ' '.join(error)
				hojaErrores.append([cuenta.codigo , ' '.join(error), cuenta.nombre])
	
		doc.save("archivosExcel/resultadoCargue.xlsx")
	
	def insertCuenta(self, cuenta):
#		inserta la cuenta en el db y se regresa el id de la cuenta
		self.cursor.execute(""" insert into financiera.cuenta_contable 
			(saldo,nombre,naturaleza,descripcion, codigo, niveL_clasificacion_cuenta_contable) 
			values ({0},LTRIM('{1}'),'{2}',LTRIM('{3}'),'{4}',{5})
			returning id;""".format(cuenta.saldo, cuenta.nombre,cuenta.naturaleza, cuenta.descripcion, cuenta.codigo, cuenta.nivel))
		cuenta.idCuenta = self.cursor.fetchone()[0]
# 		se inserta relacion padre hijo con el plan de cuentas 2
		plan_cuentas = 2
		self.cursor.execute(""" insert into financiera.estructura_cuentas (cuenta_padre, cuenta_hijo, plan_cuentas)
			values ({0},{1},{2})
			returning id""".format(cuenta.idPadre, cuenta.idCuenta, plan_cuentas))
		cuenta.idRelacion = self.cursor.fetchone()[0]
		print "Se inserta la cuenta", cuenta.codigo, ", id:",cuenta.idCuenta
		print "Se inserta relacion", cuenta.idRelacion ," con padre", cuenta.padre,"con id:", cuenta.idPadre

	def cleanIdCuenta(self,idCuentaErroneo):
		idCuentaErroneo = idCuentaErroneo.replace(" ","")
		head = idCuentaErroneo[:1]
		tail = idCuentaErroneo[1:]
		return head+"-"+tail

	def getCuentaById(self, idCuenta):
		self.cursor.execute("""
                	select id
                	from financiera.cuenta_contable
                	where codigo = trim('{0}');""".format(idCuenta))     
        	response = self.cursor.fetchone()
		if response:
			return response[0]
		return response

